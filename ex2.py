import sys
import win32api
import os
from time import sleep
import win32com.client as win32
from Tkinter import *
import tkFileDialog
import getpass
import os
from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
reload(sys)
sys.setdefaultencoding('utf-8')
wdFormatUnicodeText = 7
root = Tk()

user = getpass.getuser()
MSWord = win32.gencache.EnsureDispatch('Word.Application')
xl = win32.gencache.EnsureDispatch('Excel.Application')
myWordDoc = tkFileDialog.askopenfilename(initialdir='C:/Users/%s' % user)
myWordDoc = myWordDoc.replace("/", "\\\\")
directory = os.path.split(myWordDoc)[0]
doc = MSWord.Documents.Open(myWordDoc)
acronyms = []
txtpath = os.path.splitext(myWordDoc)[0] + '.txt'
doc.SaveAs(txtpath, wdFormatUnicodeText)
doc.Close()
#print txtpath

acronyms = []
f = file(txtpath).read()
for word in f.split():
    word = word.replace('(', '').replace(')', '').replace('[', '').replace(']', '').replace('\"', '').replace("?", '')
    word = word.strip('.!,').strip("'s'").strip("s")
    a = word[:2]
    if (a != 'MK' and a != 'ET' and a != 'BR' and a != 'MD'):
        if (a.isupper() and len(a) == 2):
            acronyms.append(word)

acronyms = sorted(list(set(acronyms)))
wb = load_workbook(filename = 'BUMED Acronym List_v3.xlsm', read_only=False, keep_vba=True, data_only=True)

ws2 = wb.get_sheet_by_name(name = "Acronyms from Deliverable")
number = len(acronyms)


N = 1
for i,e in enumerate(acronyms):
    ws2.cell(row=i+1, column=1).value = e

ws2['D1'].value = number
r = 2
excel_file_path = os.path.dirname(os.path.realpath(__file__))
excel_path = excel_file_path + "\\BUMED Acronym List_v3.xlsm"

wb.save("BUMED Acronym List_v3.xlsm")
xl.Workbooks.Open(Filename=excel_path, ReadOnly=1)
xl.Application.Run("AcronymVlookups")
xl.Application.Save
xl.Application.Quit
